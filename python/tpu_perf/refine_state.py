import os
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment


subclass = ['vision','language']
def col_f(startcol, offset):
    return chr(ord(startcol)+offset)
def row_f(startrow, offset):
    return str(startrow+offset)

sr = 0 #start row
sc = 'A' #start column
def read_config(path):
    with open(path) as f:
        return yaml.load(f, yaml.Loader)

def get_class(zoo_path):
    results = {}
    all=os.walk(zoo_path)
    for p, ds, fs in all:
        for f in fs:
            fullname = os.path.join(p,f)
            if fullname.endswith('config.yaml'):
                if zoo_path.endswith('/'):
                    subpath = fullname.replace(zoo_path, '')
                else:
                    subpath = fullname.replace(zoo_path+'/','')
                folders = subpath.split('/')
                config = read_config(fullname)
                item = dict()
                if 'name' in config:
                    item['name'] = config['name']
                    if folders[0] in subclass:
                        item['class'] = folders[0]
                        item['sub_class'] = folders[1]
                        item['folder'] = folders[2]
                    results[item['name']] = [item['class'], item['sub_class'], item['folder']]
    return results


def add_col(classtype, raw_csv):
    data = pd.read_csv(raw_csv)
    df = pd.DataFrame(data)
    type_column = []
    sub_type_column = []
    folder_column = []
    qps_column = []
    for index, row in df.iterrows():
        row['name'] = row['name'].replace("_core2", "").replace("-parallel", "")
        type_column.append(classtype[row['name']][0])
        sub_type_column.append(classtype[row['name']][1])
        folder_column.append(classtype[row['name']][2])

        model_shape = row['shape'].split("x")
        # qps = 1000 * batch / 时间(ms))
        qps = int(model_shape[0]) * 1000 / row['time(ms)'] if len(model_shape) == 4 else 'N/A'
        qps_column.append(qps)

    df.insert(loc=0, column='type', value=type_column)
    df.insert(loc=1, column='sub_type', value=sub_type_column)
    df.insert(loc=2, column='folder', value=folder_column)
    df.insert(loc=df.shape[1], column='qps', value=qps_column)
    df = df.sort_values(by=['type', 'sub_type', 'folder', 'name', 'prec'])
    df['prec'] = df['prec'].mask(df[['prec', 'name']].eq(df[['prec', 'name']].shift()).all(axis=1))
    df[['type', 'sub_type', 'folder', 'name']] = \
        df[['type', 'sub_type', 'folder', 'name']].mask(df[['type', 'sub_type', 'folder', 'name']].eq(df[['type', 'sub_type', 'folder', 'name']].shift()))
    df = df.fillna('')
    return df


def merge_cells_in_column(ws, column):
    current_value = None
    start_row = 1

    for row in range(1, ws.max_row + 2):
        cell_value = ws.cell(row=row, column=column).value

        if cell_value != '':
            if current_value is not None:
                end_row = row - 1
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)
                    ws.cell(start_row, column).alignment = Alignment(horizontal='center', vertical='center')
                start_row = row

            current_value = cell_value


def write_sheet(df, ws):
    for i, col_name in enumerate(df.columns):
        ws.cell(row=1, column=i+1, value=col_name)

    for r_idx, row in enumerate(df.iterrows(), start=1):
        for c_idx, value in enumerate(row[1], start=1):
            ws.cell(row=r_idx+1, column=c_idx, value=value)


def ajust_sheet(df, ws):
    for col in range(df.shape[1]):
        merge_cells_in_column(ws, col+1)
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = Alignment(horizontal='left', vertical='center')


def fill_sheet(df, table_name):
    wb = Workbook()
    ws = wb.active

    write_sheet(df, ws)
    ajust_sheet(df, ws)

    wb.save(table_name)
    return 


def main():
    import argparse
    parser = argparse.ArgumentParser(description='model-zoo benchmark tool')
    parser.add_argument('--stat', type=str, default='', required=True,\
                        help='the path/to/stat.csv')
    parser.add_argument('--model_zoo', type=str, default='', required=True, \
                        help='the path/to/model-zoo')
    parser.add_argument('--table_name', type=str, default='formatted_result.xlsx', \
                        help='the output file name')
    args = parser.parse_args()
    classtype = get_class(args.model_zoo)
    df = add_col(classtype, args.stat)
    fill_sheet(df, args.table_name)

    print('done')



if __name__ == '__main__':
    main()