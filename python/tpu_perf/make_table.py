from openpyxl import  Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import csv
import os
import yaml

subclass = []
netfolders = set()
target_prec_list = {'BM1684':['fp32', 'int8 1batch', 'int8 4batch', 'int8 8batch', 'int8 16batch'], \
                    'BM1684X':['fp32', 'fp16', 'int8 4batch', 'int8 8batch', 'int8 16batch'], \
                    'BM1688':['fp32', 'fp16', 'int8 4batch', 'int8 8batch', 'int8 16batch'], \
                    'CV186X':['fp32', 'fp16', 'int8 4batch', 'int8 8batch', 'int8 16batch']}
target_frequency = {'BM1684':'', 'BM1684X':'1Ghz', 'BM1688':'900Mhz', 'CV186X':'375Mhz'}

def col(startcol, offset):
    return chr(ord(startcol)+offset)
def row(startrow, offset):
    return str(startrow+offset)

sr = 2 #start row
sc = 'A' #start column
leftalign = Alignment(horizontal='left',vertical='center')

def init_table(target, tablename):

    def append_col(workbook, col_name, col_idx):
      workbook[col(sc,col_idx)+row(sr,0)] = col_name
      workbook.merge_cells(col(sc,col_idx)+row(sr,0)+':'+col(sc,col_idx)+row(sr,1))
      workbook[col(sc,col_idx)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')
      col_idx += 1
      return col_idx

    wb = Workbook()
    ws = wb.active
    ws.title = target
    col_idx = 0
    col_idx = append_col(ws, 'NetClass', col_idx)
    col_idx = append_col(ws, 'NetFolder', col_idx)
    col_idx = append_col(ws, 'ModelName', col_idx)
    col_idx = append_col(ws, 'Shape', col_idx)

    ws.title = target
    ws[col(sc,col_idx)+row(sr,0)] = target + ' @' + target_frequency[target] + ' Benchmark(qps)'
    ws.merge_cells(col(sc,col_idx)+row(sr,0)+':'+col(sc,col_idx+len(target_prec_list[target]) -1)+row(sr,0))
    ws[col(sc,col_idx)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

    for prec_idx in range(0, len(target_prec_list[target])):
      ws[col(sc,col_idx)+row(sr,1)] = target_prec_list[target][prec_idx]
      col_idx += 1

    col_idx = append_col(ws, 'Gops', col_idx)
    col_idx = append_col(ws, 'Resource', col_idx)

    wb.save(tablename)

def adjust_sheet(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
      ws=wb[sheet]
      df=pd.read_excel(filename,sheet).fillna('-')
      df.loc[len(df)]=list(df.columns)
      for column in df.columns:
        index=list(df.columns).index(column)
        letter=get_column_letter(index+1)
        collen=df[column].apply(lambda x:len(str(x).encode())).max()
        ws.column_dimensions[letter].width=collen+2

    #left side align
    if ws.title == 'BM1684X':
        endsc = 8
    else:
        endsc = 7

    colnum = 3
    for column in ws[col(sc,3)+':'+ col(sc,endsc)]:
        for rownum in range(ws.max_row):
            rownum += 1
            if (rownum > (sr+1)):
                ws[col(sc,colnum)+str(rownum)].alignment = Alignment(horizontal='left', vertical='center')
        colnum+=1

    #bind the cells with same value
    merge_value_set = [subclass, netfolders]
    merge_col_symble = ['A', 'B']
    for col_idx in [0, 1]:
      rowcnt = 1
      preval = ''
      postval = ''
      endrow = []
      first = True
      startrow = []
      for rows in ws[col(sc, col_idx)]:
          if (rows.value in merge_value_set[col_idx]) and first:
              startrow.append(rowcnt)
              first = False
              preval = rows.value
              postval = ws.cell(rowcnt+1, (ord(merge_col_symble[col_idx]) - ord('A'))+1).value

          postval = ws.cell(rowcnt+1, (ord(merge_col_symble[col_idx]) - ord('A'))+1).value
          if (preval!=postval) and (preval != ''):
              endrow.append(rowcnt)
              first = True
          rowcnt += 1
      for i in range(len(startrow)):
          ws.merge_cells(col(sc,col_idx)+str(startrow[i])+':'+col(sc,col_idx)+str(endrow[i]))
          ws[col(sc,col_idx)+str(startrow[i])].alignment = Alignment(horizontal='left', vertical='center')

    wb.save(filename)

def throughput(time, batchsize):
    fps = 1000/(float(time)/batchsize)
    return float('%.2f'%fps)

def find_class_and_folder(netname, classes):
    if classes is not None:
      for bind in classes:
        if bind[1] == netname:
          return bind[0], bind[2]
        elif netname[-6:] == '_core2' and netname[:-6] == bind[1]:
          return bind[0], bind[2]

def sort_table(input_dict):
    new_table = []
    for k in subclass:
        for item in input_dict:
            if item['class'] == k:
                new_table.append(item)
    return new_table

def get_batchsize(shape):
    step1 = shape.split(':')
    step2 = step1[0].split('x')[0]
    return int(step2)

def analyze_stat(statpath, class_type):
    bench = []
    item = dict()
    with open(statpath, 'r') as file:
      csv_file = csv.DictReader(file)
      pre_netname = ''
      new_netname = ''
      for row in csv_file:
        new_netname = row['name']
        if(('dyn' in row) and (row['dyn'] == 'TRUE')):
          continue
        if pre_netname != new_netname:
          tmp = item.copy()
          if tmp!= {}:
            bench.append(tmp)
          item.clear()
          item={'class':'','name':'','shape':'','fp32':'N/A','fp16':'N/A','int8-1b':'N/A', \
                'int8-4b':'N/A','int8-8b':'N/A','int8-16b':'N/A','gops':'N/A'}
          shape = row['shape'].split('x')[1:]
          dims = '*'.join(shape)
          item['class'], item['net_folder'] = find_class_and_folder(row['name'], class_type)
          item['name'] = row['name']
          item['shape'] = dims
          item['gops'] = row['gops']
          pre_netname = new_netname
          time = 'time(ms)'
          if row['prec']=='FP32':
            item['fp32'] = throughput(row[time], get_batchsize(row['shape']))
          elif ((row['prec'] == 'FP16') or (row['prec'] == 'BF16')):
              item['fp16'] = throughput(row[time], get_batchsize(row['shape']))
          else:
            if(row['shape'].split('x')[0]=='1'):
              item['int8-1b'] = throughput(row[time], 1)
            elif(row['shape'].split('x')[0]=='4'):
              item['int8-4b'] = throughput(row[time], 4)
            elif(row['shape'].split('x')[0]=='8'):
              item['int8-8b'] = throughput(row[time], 8)
            elif(row['shape'].split('x')[0]=='16'):
              item['int8-16b'] = throughput(row[time], 16)
        else:
          if row['prec']=='FP32':
            item['fp32'] = throughput(row[time], get_batchsize(row['shape']))
          elif ((row['prec'] == 'FP16') or (row['prec'] == 'BF16')):
            item['fp16'] = throughput(row[time], get_batchsize(row['shape']))
          else:
            if(row['shape'].split('x')[0]=='1'):
              item['int8-1b'] = throughput(row[time], 1)
            elif(row['shape'].split('x')[0]=='4'):
              item['int8-4b'] = throughput(row[time], 4)
            elif(row['shape'].split('x')[0]=='8'):
              item['int8-8b'] = throughput(row[time], 8)
            elif(row['shape'].split('x')[0]=='16'):
              item['int8-16b'] = throughput(row[time], 16)

      tmp = item.copy()
      bench.append(tmp)
      return sort_table(bench)

def fill_table(bench, tablename, target):
    wb = load_workbook(tablename)
    ws = wb.active
    #ws.append(['name', 'shape', 'fp32', 'fp16','int8-1batch','int8-4batch','int8-8batch','int8-16batch'])
    for item in bench:
        if target=='BM1684':
            ws.append([item['class'], item['net_folder'], item['name'], item['shape'], item['fp32'], \
                     item['int8-1b'],item['int8-4b'],item['int8-8b'], \
                     item['int8-16b'], item['gops']])
        else:
            ws.append([item['class'], item['net_folder'], item['name'], item['shape'], item['fp32'],item['fp16'], \
                     item['int8-1b'],item['int8-4b'],item['int8-8b'], \
                     item['int8-16b'], item['gops']])

    wb.save(tablename)

def read_config(path):
    '''
    fn = os.path.join(path, 'config.yaml')
    if not os.path.exists(fn):
        print(f'No config in {path}')
        return
    '''
    with open(path) as f:
        return yaml.load(f, yaml.Loader)

def get_class(zoo_path):
    results = []
    all=os.walk(zoo_path)
    for p, ds, fs in all:
        for f in fs:
            fullname = os.path.join(p,f)
            if fullname.endswith('config.yaml'):
                if zoo_path.endswith('/'):
                    subpath = fullname.replace(zoo_path, '')
                else:
                    subpath = fullname.replace(zoo_path+'/','')
                folders = subpath.split('/')
                if len(folders) > 1: #in sub folder
                    if folders[1] not in subclass:
                        subclass.append(folders[1])
                config = read_config(fullname)
                item = dict()
                if 'name' in config:
                    item['name'] = config['name']
                    if folders[1] in subclass:
                        item['class'] = folders[1]
                        item['net_folder'] = folders[2]
                        netfolders.add(folders[2])
                    #item['gops'] = config['gops']
                    results.append((item['class'],item['name'], item['net_folder']))
    subclass.sort()
    return results

def main():
    import argparse
    parser = argparse.ArgumentParser(description='model-zoo benchmark tool')
    parser.add_argument('--stat', type=str, default='', required=True,\
                        help='the path/to/stat.csv')
    parser.add_argument('--target', type=str, default='BM1684X', \
                        help='the device type, BM1684, BM1684X, BM1688, CV186X')
    parser.add_argument('--model_zoo', type=str, default='', required=True, \
                        help='the path/to/model-zoo')
    parser.add_argument('--table_name', type=str, default='formatted_result.xlsx', \
                        help='the output file name')
    args = parser.parse_args()
    classtype = get_class(args.model_zoo)
    init_table(args.target, args.table_name)
    output = analyze_stat(args.stat, classtype)
    fill_table(output, args.table_name, args.target)
    adjust_sheet(args.table_name)


if __name__ == '__main__':
    main()
